from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
import os
from typing import List, Optional
from pathlib import Path
import secrets  # <-- Убедитесь, что этот импорт есть

from app import models, schemas
from app.dependencies import get_db
from app.routers.auth import get_current_user

# Импорты для PDF и Excel
from reportlab.platypus import Table, TableStyle, Paragraph, Image as PlatypusImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

router = APIRouter()

BASE_DIR = Path(__file__).resolve().parent.parent.parent
INVOICE_DIR = BASE_DIR / "invoices"
FONT_PATH = BASE_DIR / "DejaVuSans.ttf"
pdfmetrics.registerFont(TTFont('DejaVuSans', str(FONT_PATH)))
pdfmetrics.registerFont(TTFont('DejaVuSans-Oblique', str(BASE_DIR / "DejaVuSans-Oblique.ttf")))


def generate_invoice_files(order: models.Order, db: Session):
    date_str = order.created_at.strftime("%Y-%m")
    folder_path = INVOICE_DIR / date_str
    os.makedirs(folder_path, exist_ok=True)

    file_base_name = f"order_{order.order_number}"
    pdf_path = folder_path / f"{file_base_name}.pdf"
    xlsx_path = folder_path / f"{file_base_name}.xlsx"

    subtotal = sum(item.quantity * item.price_per_item for item in order.items)
    discount_amount = (subtotal * order.discount_percent) / 100
    final_total = subtotal - discount_amount

    # --- Генерация PDF ---
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    c.setFont('DejaVuSans', 14)
    c.drawString(72, height - 50, f"Расходная накладная № {order.order_number}")
    c.setFont('DejaVuSans', 11)
    c.drawString(72, height - 70, f"Дата: {order.created_at.strftime('%d.%m.%Y %H:%M')}")
    c.drawString(72, height - 90, f"Клиент: {order.customer_name}, Телефон: {order.customer_phone}")
    
    # --- ИЗМЕНЕНИЕ НАЧАЛО ---
    # Определяем стили для ячеек
    styles = getSampleStyleSheet()
    # Стиль для обычного текста в ячейке. wordWrap = 'CJK' позволяет переносить длинные строки без пробелов.
    styleN = ParagraphStyle(name='Normal', fontName='DejaVuSans', fontSize=9, wordWrap='CJK', alignment=1)
    # Стиль для комментариев
    styleComment = ParagraphStyle(name='Comment', parent=styleN, fontName='DejaVuSans-Oblique', textColor=colors.grey)
    # Стиль для правого выравнивания (цены, количество)
    styleR = ParagraphStyle(name='Right', parent=styleN, alignment=2)

    # Заголовки таблицы оборачиваем в параграфы для единообразия
    headers = [Paragraph(h, styleN) for h in ['№', 'Артикул', 'Фото', 'Наименование товара', 'Кол-во', 'Цена', 'Сумма']]
    data = [headers]
    
    for i, item in enumerate(order.items):
        product = db.query(models.Product).filter(models.Product.id == item.product_id).first()
        item_sum = item.quantity * item.price_per_item
        product_name = product.name if product else "Товар не найден"
        
        image_cell = ''
        if product and product.image_url:
            image_path = BASE_DIR / product.image_url.lstrip('/')
            if image_path.exists():
                img = PlatypusImage(image_path, width=2*cm, height=2*cm)
                image_cell = img
        
        # Оборачиваем содержимое каждой ячейки в Paragraph
        name_cell_content = [Paragraph(product_name, styleN)]
        if item.comment:
            name_cell_content.append(Paragraph(item.comment, styleComment))

        row = [
            Paragraph(str(i + 1), styleN),
            Paragraph(product.sku if product else "N/A", styleN), # Артикул теперь тоже Paragraph
            image_cell,
            name_cell_content,
            Paragraph(str(item.quantity), styleR),
            Paragraph(f"{item.price_per_item:.2f}", styleR),
            Paragraph(f"{item_sum:.2f}", styleR)
        ]
        data.append(row)
    # --- ИЗМЕНЕНИЕ КОНЕЦ ---

    col_widths = [1*cm, 2.2*cm, 2.2*cm, 5.1*cm, 1.5*cm, 2.5*cm, 2.5*cm] # Слегка изменили ширину Артикула
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
        ('GRID', (0,0), (-1,-1), 1, colors.grey),
    ]))
    table_height = table.wrapOn(c, width - 144, height)[1]
    table.drawOn(c, 72, height - 100 - table_height)
    
    y_pos = height - 110 - table_height - 20
    c.setFont('DejaVuSans', 10)
    c.drawRightString(width - 72, y_pos, f"Сумма: {subtotal:.2f}")
    if order.discount_percent > 0:
        y_pos -= 15
        c.drawRightString(width - 72, y_pos, f"Скидка ({order.discount_percent:.1f}%): -{discount_amount:.2f}")
    
    c.setFont('DejaVuSans', 12)
    c.drawRightString(width - 72, y_pos - 20, f"Итого к оплате: {final_total:.2f}") # Скорректировали отступ
    c.save()

    # --- Генерация Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"Накладная {order.order_number}"
    ws.append(["Расходная накладная №", order.order_number])
    ws.append(["Дата:", order.created_at.strftime('%d.%m.%Y %H:%M')])
    ws.append(["Клиент:", order.customer_name, "Телефон:", order.customer_phone])
    ws.append([])
    headers_row = ["№", "Артикул", "Наименование", "Кол-во", "Цена за ед.", "Сумма"]
    ws.append(headers_row)
    start_table_row = ws.max_row
    for i, item in enumerate(order.items):
        product = db.query(models.Product).filter(models.Product.id == item.product_id).first()
        item_sum = item.quantity * item.price_per_item
        product_name = product.name if product else "Товар не найден"
        if item.comment:
            product_name += f"\n(Комментарий: {item.comment})"
        ws.append([ i + 1, product.sku if product else "N/A", product_name, item.quantity, item.price_per_item, item_sum ])
    ws.append([])
    ws.append(["", "", "", "", "Сумма:", subtotal])
    if order.discount_percent > 0:
        ws.append(["", "", "", "", f"Скидка ({order.discount_percent}%):", f"-{discount_amount}"])
    ws.append(["", "", "", "", "Итого к оплате:", final_total])
    thin_border_side = Side(border_style="thin", color="000000")
    full_border = Border(top=thin_border_side, left=thin_border_side, right=thin_border_side, bottom=thin_border_side)
    for row in ws.iter_rows(min_row=start_table_row, max_row=ws.max_row, min_col=1, max_col=len(headers_row)):
        for cell in row:
            cell.border = full_border
            if cell.column == 3: cell.alignment = Alignment(wrap_text=True, vertical='top')
            elif cell.column > 3: cell.alignment = Alignment(horizontal='right', vertical='center')
    wb.save(str(xlsx_path))

    # --- ИЗМЕНЕНИЕ: Возвращаем ВЕБ-ПУТЬ ---
    pdf_url_path = f"/invoices/{date_str}/{file_base_name}.pdf"
    xlsx_url_path = f"/invoices/{date_str}/{file_base_name}.xlsx"
    
    return pdf_url_path, xlsx_url_path


@router.post("/", response_model=schemas.Order)
def create_order(order_data: schemas.OrderCreate, db: Session = Depends(get_db)):
    
    # --- ИЗМЕНЕНИЕ НАЧАЛО ---
    # 1. Сразу получаем настройку игнорирования остатков
    ignore_stock_setting = db.query(models.AppSettings).filter(models.AppSettings.key == "ignore_stock_limits").first()
    ignore_stock = ignore_stock_setting.value == "true" if ignore_stock_setting else False

    # 2. Сначала проверяем все товары в корзине, ПЕРЕД созданием заказа
    for item_data in order_data.items:
        product = db.query(models.Product).filter(models.Product.id == item_data.product_id).first()
        if not product:
            # Если хоть один товар не найден, заказ не создаем
            raise HTTPException(status_code=404, detail=f"Товар с ID {item_data.product_id} не найден.")
        
        # 3. Выполняем проверку остатков, если настройка выключена
        if not ignore_stock and product.stock < item_data.quantity:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Недостаточно товара '{product.name}' на складе. Доступно: {product.stock}, в заказе: {item_data.quantity}."
            )
    # --- ИЗМЕНЕНИЕ КОНЕЦ ---

    # Если все проверки пройдены, создаем заказ как обычно
    new_order = models.Order(
        customer_name=order_data.customer_name,
        customer_phone=order_data.customer_phone,
        customer_comment=order_data.customer_comment,
        private_key=secrets.token_urlsafe(16)
    )
    db.add(new_order)
    db.commit()
    db.refresh(new_order)

    new_order.order_number = f"{datetime.now().strftime('%Y-%m')}-{new_order.id:05d}"
    
    for item_data in order_data.items:
        product = db.query(models.Product).filter(models.Product.id == item_data.product_id).first()
        # Повторная проверка на продукт здесь уже для безопасности, но основная логика выше
        if not product:
            db.rollback()
            raise HTTPException(status_code=404, detail=f"Product with id {item_data.product_id} not found")
        
        order_item = models.OrderItem(
            order_id=new_order.id, 
            product_id=item_data.product_id, 
            quantity=item_data.quantity, 
            price_per_item=product.price, 
            comment=item_data.comment
        )
        db.add(order_item)
    
    db.commit()
    db.refresh(new_order)
    
    pdf_path, xlsx_path = generate_invoice_files(new_order, db)
    new_order.invoice_path_pdf = pdf_path
    new_order.invoice_path_xlsx = xlsx_path
    db.commit()
    db.refresh(new_order)
    
    return new_order


@router.get("/", response_model=List[schemas.Order])
def get_orders(db: Session = Depends(get_db)):
    orders = db.query(models.Order).order_by(models.Order.id.desc()).all()
    return orders